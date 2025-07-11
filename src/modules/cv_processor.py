# modules/cv_processor.py

import os  # xử lý tương tác với hệ thống file và biến môi trường
import re  # xử lý biểu thức chính quy
import json  # parse và dump JSON
import time  # xử lý thời gian và sleep retry
import logging  # ghi log
from datetime import datetime, date  # định dạng thời gian hiển thị và lọc
from typing import List, Dict, Optional, Callable  # khai báo kiểu

import pandas as pd  # xử lý DataFrame
pd.set_option("display.max_colwidth", None)  # hiển thị đầy đủ nội dung các cột
import docx  # đọc file .docx
from openpyxl.styles import Font, PatternFill  # định dạng Excel

# --- Thiết lập logger cho module ---
logger = logging.getLogger(__name__)  # lấy logger theo tên module
logger.setLevel(logging.INFO)  # mức độ log tối thiểu INFO
fmt = logging.Formatter("%(asctime)s - %(levelname)s - %(message)s")  # định dạng log
# handler xuất log ra console
stream_h = logging.StreamHandler()  
stream_h.setFormatter(fmt)
logger.addHandler(stream_h)
# handler xuất log ra file trong thư mục log
from .config import LOG_DIR
LOG_DIR.mkdir(parents=True, exist_ok=True)
file_h = logging.FileHandler(LOG_DIR / "cv_processor.log", encoding="utf-8")
file_h.setFormatter(fmt)
logger.addHandler(file_h)

# --- Cấu hình extractor PDF: pdfminer, PyPDF2 hoặc PyMuPDF ---
_PDF_EX: Optional[str]
try:
    from pdfminer.high_level import extract_text
    _PDF_EX = "pdfminer"  # ưu tiên pdfminer nếu cài đặt
except ImportError:
    try:
        import PyPDF2  # noqa: F401
        _PDF_EX = "pypdf2"  # fallback sang PyPDF2
    except ImportError:
        try:
            import fitz  # noqa: F401  # PyMuPDF
            _PDF_EX = "pymupdf"  # fallback sang PyMuPDF
        except ImportError:
            _PDF_EX = None  # không có thư viện PDF nào

from .llm_client import LLMClient  # client LLM mặc định
from typing import Union
# Support both LLMClient and DynamicLLMClient
from .config import (
    ATTACHMENT_DIR,
    OUTPUT_CSV,
    OUTPUT_EXCEL,
    EMAIL_UNSEEN_ONLY,
)
from .sent_time_store import load_sent_times
from .prompts import CV_EXTRACTION_PROMPT  # prompt LLM để trích xuất CV

def format_sent_time_display(ts: str) -> str:
    """Định dạng thời gian ISO sang dạng dễ đọc hơn."""
    if not ts:
        return ""
    try:
        dt = datetime.fromisoformat(ts.replace("Z", "+00:00"))
        dt = dt.astimezone()
        hour = dt.strftime("%I").lstrip("0") or "0"
        minute = dt.strftime("%M")
        ampm = dt.strftime("%p").lower()
        day = dt.strftime("%d").lstrip("0")
        month = dt.strftime("%m").lstrip("0")
        year = dt.strftime("%Y")
        return f"{hour}:{minute} {ampm} {day}/{month}/{year}"
    except Exception:
        return ts

class CVProcessor:
    """
    Lớp xử lý file CV: đọc text, gọi LLM hoặc regex fallback, trả về DataFrame
    """
    def __init__(self, fetcher: Optional[object] = None, llm_client = None):
        """Khởi tạo: cấp fetcher (đọc email) và LLM client"""
        self.fetcher = fetcher  # đối tượng có method fetch_cv_attachments()
        self.llm_client = llm_client or LLMClient()  # client LLM mặc định

    def _extract_pdf(self, path: str) -> str:
        """
        Đọc text từ file PDF bằng thư viện tương ứng
        Trả về chuỗi rỗng nếu không có library
        """
        if _PDF_EX == "pdfminer":
            return extract_text(path)
        elif _PDF_EX == "pypdf2":
            from PyPDF2 import PdfReader
            txt = ""
            for p in PdfReader(path).pages:
                txt += p.extract_text() or ""
            return txt
        elif _PDF_EX == "pymupdf":
            import fitz
            doc = fitz.open(path)
            txt = "".join(page.get_text() for page in doc)
            doc.close()
            return txt
        logger.error("❌ Không có thư viện PDF phù hợp để trích xuất text.")
        return ""

    def extract_text(self, path: str) -> str:
        """
        Đọc văn bản từ file PDF hoặc DOCX
        Trả về chuỗi text, log cảnh báo nếu định dạng không hỗ trợ
        """
        ext = os.path.splitext(path)[1].lower()  # lấy phần mở rộng
        try:
            if ext == ".pdf":
                return self._extract_pdf(path)
            if ext == ".docx":
                doc = docx.Document(path)
                return "\n".join(p.text for p in doc.paragraphs)
            logger.warning(f"⚠️ Định dạng không hỗ trợ: {path}")
        except Exception as e:
            logger.error(f"Lỗi khi đọc file {path}: {e}")
        return ""

    def extract_info_with_llm(self, text: str) -> Dict:
        """
        Enhanced LLM extraction with better error handling and retry logic
        """
        if not text.strip():
            logger.warning("Text input is empty for LLM extraction")
            return {}

        max_retries = 3
        base_delay = 1  # seconds
        
        for attempt in range(1, max_retries + 1):
            try:
                logger.info(f"LLM extraction attempt {attempt}/{max_retries}")
                
                # Create enhanced prompt with text length info
                text_length = len(text)
                if text_length > 8000:  # Truncate very long text
                    text = text[:8000] + "...[text truncated]"
                    logger.info(f"Text truncated from {text_length} to {len(text)} chars")
                
                # Generate response with timeout
                resp = self.llm_client.generate_content([CV_EXTRACTION_PROMPT, text])
                
                if not resp or not resp.strip():
                    raise ValueError(f"Empty response from LLM on attempt {attempt}")
                
                # Enhanced JSON extraction with multiple patterns
                json_data = self._extract_json_from_response(resp)
                
                if json_data:
                    logger.info(f"✅ LLM extraction successful on attempt {attempt}")
                    return json_data
                else:
                    raise ValueError(f"No valid JSON found in LLM response on attempt {attempt}")
                    
            except Exception as e:
                logger.warning(f"❌ LLM attempt {attempt} failed: {e}")
                
                # Check for quota/rate limit errors
                error_msg = str(e).lower()
                if any(code in error_msg for code in ("quota", "429", "resource_exhausted", "rate limit")):
                    if attempt < max_retries:
                        delay = base_delay * (2 ** attempt)  # Exponential backoff
                        logger.warning(f"Quota/rate limit hit, retrying in {delay} seconds...")
                        time.sleep(delay)
                        continue
                
                if attempt == max_retries:
                    logger.error(f"All {max_retries} LLM attempts failed. Falling back to regex.")
                    return self._fallback_regex(text)
        
        # Fallback to regex if all attempts fail
        return self._fallback_regex(text)

    def _extract_json_from_response(self, response: str) -> Optional[Dict]:
        """Extract JSON from LLM response with multiple patterns"""
        try:
            # Pattern 1: JSON in code blocks
            json_match = re.search(r'```(?:json)?\s*(\{.*?\})\s*```', response, re.DOTALL)
            if json_match:
                return json.loads(json_match.group(1))
            
            # Pattern 2: Direct JSON
            json_match = re.search(r'\{.*\}', response, re.DOTALL)
            if json_match:
                return json.loads(json_match.group(0))
            
            # Pattern 3: Try to parse entire response as JSON
            return json.loads(response.strip())
            
        except json.JSONDecodeError as e:
            logger.warning(f"JSON parsing failed: {e}")
            return None
        except Exception as e:
            logger.error(f"Unexpected error in JSON extraction: {e}")
            return None

    def _fallback_regex(self, text: str) -> Dict:
        """
        Dùng regex đơn giản để trích xuất các trường: tên, tuổi, email, điện thoại, học vấn, kinh nghiệm, địa chỉ, kỹ năng
        Trả về dict với các key tương ứng
        """
        patterns = {
            "ten": r"(?:(?:Họ tên|Tên)[:\-\s]+)([^\n]+)",
            "tuoi": r"(?:(?:Tuổi|Age)[:\-\s]+)(\d{1,3})",
            "email": r"([a-zA-Z0-9_.+-]+@[a-zA-Z0-9-]+\.[a-zA-Z0-9-.]+)",
            "dien_thoai": r"(\+?\d[\d\-\s]{7,}\d)",
            "vi_tri": r"(?:(?:Vị trí|Position)[:\-\s]+)([^\n]+)",
            "hoc_van": r"(?:(?:Học vấn|Education)[:\-\s]+)([^\n]+)",
            "kinh_nghiem": r"(?:(?:Kinh nghiệm|Experience)[:\-\s]+)([^\n]+)",
            "dia_chi": r"(?:(?:Địa chỉ|Address)[:\-\s]+)([^\n]+)",
            "ky_nang": r"(?:(?:Kỹ năng|Skills?)[:\-\s]+)([^\n]+)",
        }
        info: Dict[str, str] = {}
        for k, p in patterns.items():
            m = re.search(p, text, re.IGNORECASE)
            info[k] = m.group(1).strip() if m else ""
        return info

    def process(
        self,
        unseen_only: bool | None = None,
        since: date | None = None,
        before: date | None = None,
        from_time: datetime | None = None,
        to_time: datetime | None = None,
        progress_callback: Optional[Callable[[int, str], None]] = None,
        ignore_last_uid: bool = False,
    ) -> pd.DataFrame:
        """
        Tìm tất cả file CV (fetcher hoặc thư mục attachments), trích xuất info, trả về DataFrame
        """
        # fetch từ email nếu có fetcher
        if self.fetcher:
            unseen = unseen_only if unseen_only is not None else EMAIL_UNSEEN_ONLY
            fetch_kwargs = {
                "since": since,
                "before": before,
                "unseen_only": unseen,
            }
            import inspect
            sig = inspect.signature(self.fetcher.fetch_cv_attachments)
            if "ignore_last_uid" in sig.parameters:
                fetch_kwargs["ignore_last_uid"] = ignore_last_uid
            files: List[str] = self.fetcher.fetch_cv_attachments(**fetch_kwargs)
        else:
            files = []

        sent_map = {
            os.path.join(ATTACHMENT_DIR, fname): ts
            for fname, ts in load_sent_times().items()
        }
        if self.fetcher:
            sent_map.update(dict(getattr(self.fetcher, "last_fetch_info", [])))
        if not files:
            logger.info("🔍 dò thư mục attachments...")
            files = [
                os.path.join(ATTACHMENT_DIR, f)
                for f in os.listdir(ATTACHMENT_DIR)
                if f.lower().endswith((".pdf", ".docx"))
            ]

        if from_time or to_time:
            def _in_range(p: str) -> bool:
                ts = sent_map.get(p, "")
                if not ts:
                    return False
                try:
                    dt = datetime.fromisoformat(ts.replace("Z", "+00:00"))
                except Exception:
                    return False
                if from_time and dt < from_time:
                    return False
                if to_time and dt > to_time:
                    return False
                return True

            files = [f for f in files if _in_range(f)]

        total_files = len(files)
        if progress_callback:
            progress_callback(0, f"Bắt đầu xử lý {total_files} file...")

        if not files:
            logger.info("ℹ️ Không có file CV nào trong thư mục.")
            return pd.DataFrame()  # trả về DataFrame rỗng nếu không có file

        rows: List[Dict[str, str]] = []
        for idx, path in enumerate(files):
            txt = self.extract_text(path)  # đọc text file
            info = self.extract_info_with_llm(txt) or {}
            # gom thông tin vào dict
            sent_time = sent_map.get(path, "")
            sent_time = sent_time if sent_time is not None else ""
            rows.append({
                "Thời gian nhận": sent_time,
                "Nguồn": os.path.basename(path),
                "Vị trí": info.get("vi_tri", ""),
                "Họ tên": info.get("ten", ""),
                "Tuổi": info.get("tuoi", ""),
                "Email": info.get("email", ""),
                "Điện thoại": info.get("dien_thoai", ""),
                "Địa chỉ": info.get("dia_chi", ""),
                "Học vấn": info.get("hoc_van", ""),
                "Kinh nghiệm": info.get("kinh_nghiem", ""),
                "Kỹ năng": info.get("ky_nang", ""),
            })

            if progress_callback:
                percentage = ((idx + 1) / total_files) * 100 if total_files > 0 else 100
                progress_callback(idx + 1, f"Đang xử lý {os.path.basename(path)} ({percentage:.1f}%)")

        df = pd.DataFrame(rows, columns=[
            "Thời gian nhận",
            "Nguồn",
            "Vị trí",
            "Họ tên",
            "Tuổi",
            "Email",
            "Điện thoại",
            "Địa chỉ",
            "Học vấn",
            "Kinh nghiệm",
            "Kỹ năng",
        ])  # tạo DataFrame từ list dict với thứ tự cột cố định
        if hasattr(df, "sort_values"):
            df.sort_values("Thời gian nhận", ascending=False, inplace=True)
            df["Thời gian nhận"] = df["Thời gian nhận"].apply(format_sent_time_display)
        else:
            df.sort(key=lambda r: r.get("Thời gian nhận", ""), reverse=True)
            for row in df:
                row["Thời gian nhận"] = format_sent_time_display(row.get("Thời gian nhận", ""))

        if progress_callback:
            progress_callback(total_files, f"✅ Hoàn tất xử lý {total_files} file")

        return df  # trả về kết quả

    def save_to_csv(self, df: pd.DataFrame, output: str = OUTPUT_CSV):
        """
        Ghi đè file CSV mỗi lần chạy; nếu muốn append, có thể chuyển mode và header
        """
        df.to_csv(output, index=False, encoding="utf-8-sig")  # lưu file
        logger.info(f"✅ Đã lưu {len(df)} hồ sơ vào {output}")

    def save_to_excel(self, df: pd.DataFrame, output: str = OUTPUT_EXCEL) -> None:
        """Ghi DataFrame ra file Excel với định dạng đẹp và hyperlink"""
        df = df.copy()
        with pd.ExcelWriter(output, engine="openpyxl") as writer:
            df.to_excel(writer, index=False, sheet_name="CVs")
            wb = writer.book
            ws = writer.sheets["CVs"]

            # Định dạng header
            header_font = Font(bold=True)
            header_fill = PatternFill("solid", fgColor="FFFFCC")
            for cell in ws[1]:
                cell.font = header_font
                cell.fill = header_fill

            ws.freeze_panes = "A2"

            # Auto-width cho các cột
            for column_cells in ws.columns:
                length = max(len(str(cell.value)) if cell.value is not None else 0 for cell in column_cells)
                ws.column_dimensions[column_cells[0].column_letter].width = min(length + 2, 50)

            # Tạo hyperlink cho cột "Nguồn"
            if "Nguồn" in df.columns:
                col_idx = list(df.columns).index("Nguồn") + 1
                for row in range(2, len(df) + 2):
                    fname = ws.cell(row=row, column=col_idx).value
                    path = (ATTACHMENT_DIR / str(fname)).resolve()
                    ws.cell(row=row, column=col_idx).value = f'=HYPERLINK("{path}", "{fname}")'

        logger.info(f"✅ Đã lưu {len(df)} hồ sơ vào {output}")
